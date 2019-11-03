#!/usr/bin/env python
#-*- coding:utf-8 -*-
# datetime:2019/10/21 22:07
# software: PyCharm
__author__ = '刘影'
import xlwt
import xlrd
import openpyxl
import os
from common.log import Log
log = Log()
class ExcelOperate:
    #向excel中写入数据
    '''
     @ param Result
     @ param RowNum
     @ param ColNum
    @ param Path
    @ param SheetName
    '''
    def setCellData(self,result,rownum,colnum,path,sheetname="sheet1"):
        excel = xlwt.Workbook()
        sheet = excel.add_sheet(sheetname)
        sheet.write(rownum,colnum,result)
        excel.save(path)
        log.info("写入EXCEL单元格成功")
#读取excel单元格的值
    def getCellData(self,path,sheetname,rownum,colnum):
        excelbook = xlrd.open_workbook(path)
        excelsheet = excelbook.sheet_by_name(sheetname)
        result = excelsheet.cell(rownum,colnum)
        log.info("读取EXCEL单元格,单元格值为{0}".format(result))
        return result
    #使用openpyxl读取单元格值
    def readCell(self,path,sheetname,rownum,cellnum):
        wb = openpyxl.load_workbook(path)
        s1 = wb[sheetname]
        result = s1.cell(rownum,cellnum).value
        log.info("使用openpyxl读取EXCEL单元格,单元格值为[{0}]".format(result))
        return result
    #使用openpyxl写excel
    def writeCell(self,path,sheetname,rownum,cellnum,value):
        if not os.path.exists(path):
            wb = openpyxl.Workbook()
            sh = wb.create_sheet(sheetname)
        else:
            wb = openpyxl.load_workbook(path)
            sh = wb[sheetname]
        sh.cell(rownum, cellnum).value = value
        wb.save(path)
        log.info("使用openpyxl写入EXCEL单元格成功")
